import os
import uuid
import hashlib
import math
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from decimal import Decimal
import pandas as pd
from openpyxl import load_workbook


def sanitize_for_json(obj: Any) -> Any:
    """Reemplaza NaN/Inf de Python con None para serialización JSON válida."""
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    elif isinstance(obj, dict):
        return {k: sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_for_json(v) for v in obj]
    return obj

from app.domain.entities.models import FileUploadStatus


class ExcelProcessingError(Exception):
    def __init__(self, message: str, row: Optional[int] = None, column: Optional[str] = None):
        self.message = message
        self.row = row
        self.column = column
        super().__init__(self.message)


class DataCleaner:
    def __init__(self, df: pd.DataFrame, file_upload_id: uuid.UUID):
        self.df = df.copy()
        self.file_upload_id = file_upload_id
        self.errors: List[Dict[str, Any]] = []
        self.warnings: List[str] = []
        self.rows_processed = 0

    def clean(self) -> pd.DataFrame:
        self._standardize_columns()
        self._handle_missing_values()
        self._validate_data_types()
        self._detect_outliers()
        self._normalize_strings()
        return self.df

    def _standardize_columns(self):
        column_mapping = {}
        for col in self.df.columns:
            normalized = col.lower().strip().replace(" ", "_").replace("ñ", "n")
            column_mapping[col] = normalized
        
        self.df = self.df.rename(columns=column_mapping)
        
        required_cols = ["timestamp", "value"]
        if not all(col in self.df.columns for col in required_cols):
            self.warnings.append(f"Columnas esperadas: {required_cols}")

    def _handle_missing_values(self):
        for col in self.df.columns:
            null_count = self.df[col].isnull().sum()
            if null_count > 0:
                if col in ["timestamp", "value", "sensor_id", "sensor_code"]:
                    self.df = self.df.dropna(subset=[col])
                    self.errors.append({
                        "type": "missing_values",
                        "column": col,
                        "count": int(null_count),
                        "action": "dropped"
                    })
                else:
                    if self.df[col].dtype in ["float64", "int64"]:
                        self.df[col] = self.df[col].fillna(0)
                    else:
                        self.df[col] = self.df[col].fillna("desconocido")

    def _validate_data_types(self):
        if "timestamp" in self.df.columns:
            try:
                self.df["timestamp"] = pd.to_datetime(self.df["timestamp"], errors="coerce")
                invalid_dates = self.df["timestamp"].isnull().sum()
                if invalid_dates > 0:
                    self.errors.append({
                        "type": "invalid_dates",
                        "count": int(invalid_dates),
                        "action": "set_to_null"
                    })
            except Exception as e:
                self.errors.append({"type": "date_conversion_error", "message": str(e)})

        if "value" in self.df.columns:
            self.df["value"] = pd.to_numeric(self.df["value"], errors="coerce")
            invalid_values = self.df["value"].isnull().sum()
            if invalid_values > 0:
                self.errors.append({
                    "type": "invalid_values",
                    "count": int(invalid_values),
                    "action": "set_to_null"
                })

    def _detect_outliers(self):
        if "value" in self.df.columns:
            Q1 = self.df["value"].quantile(0.25)
            Q3 = self.df["value"].quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 3 * IQR
            upper_bound = Q3 + 3 * IQR
            
            outliers = (
                (self.df["value"] < lower_bound) | 
                (self.df["value"] > upper_bound)
            )
            outlier_count = outliers.sum()
            
            if outlier_count > 0:
                self.df.loc[outliers, "is_outlier"] = True
                self.warnings.append(f"Detectados {int(outlier_count)} valores atípicos (>3 IQR)")

    def _normalize_strings(self):
        string_cols = self.df.select_dtypes(include=["object"]).columns
        for col in string_cols:
            if col not in ["is_outlier"]:
                self.df[col] = self.df[col].astype(str).str.strip().str.lower()

    def get_summary(self) -> Dict[str, Any]:
        return {
            "total_rows": len(self.df),
            "rows_processed": self.rows_processed,
            "errors": self.errors,
            "warnings": self.warnings,
            "columns": list(self.df.columns)
        }


class ExcelParser:
    SHEET_MAPPINGS = {
        'sensores': ['sensores', 'sensor', 'medidores', 'contadores', 'registros'],
        'consumos': ['gesmed', 'volumenes', 'consumos', 'consumo', 'volumen', 'lecturas', 'costes', 'explotacion'],
        'ubicaciones': ['ubicaciones', 'ubicacion', 'ubic', 'locations', 'location', 'desaladoras']
    }
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.metadata: Dict[str, Any] = {}
        self.is_xls = file_path.lower().endswith('.xls') and not file_path.lower().endswith('.xlsx')
        self.sheet_map: Dict[str, str] = {}

    def _find_sheet(self, target_type: str) -> str | None:
        for target_name in self.SHEET_MAPPINGS.get(target_type, []):
            for sheet_name in self.sheet_map.values():
                if target_name.lower() in sheet_name.lower():
                    return sheet_name
        return None

    def _find_all_sheets(self, target_type: str) -> List[str]:
        """Retorna TODAS las hojas que coincidan con el tipo dado (sin duplicados)."""
        found = []
        for target_name in self.SHEET_MAPPINGS.get(target_type, []):
            for sheet_name in self.sheet_map.values():
                if target_name.lower() in sheet_name.lower() and sheet_name not in found:
                    found.append(sheet_name)
        return found

    def validate_structure(self) -> Tuple[bool, str]:
        try:
            if self.is_xls:
                import xlrd
                self.workbook = xlrd.open_workbook(self.file_path)
                sheet_names = self.workbook.sheet_names()
            else:
                self.workbook = load_workbook(self.file_path, read_only=True)
                sheet_names = self.workbook.sheetnames
            
            if not sheet_names:
                return False, "El archivo no contiene hojas"
            
            self.sheet_map = {name.lower(): name for name in sheet_names}
            self.metadata["sheets"] = sheet_names
            self.metadata["sheet_count"] = len(sheet_names)
            
            found_sensors = self._find_sheet('sensores')
            found_consumos = self._find_sheet('consumos')
            found_ubicaciones = self._find_sheet('ubicaciones')
            
            if not (found_sensors or found_consumos or found_ubicaciones):
                return False, f"No se encontraron hojas reconocidas. Hojas disponibles: {sheet_names}"
            
            return True, "Estructura válida"
        except Exception as e:
            return False, f"Error al validar estructura: {str(e)}"

    def _get_sheet_name(self, target_type: str) -> str:
        sheet = self._find_sheet(target_type)
        if sheet:
            return sheet
        return target_type

    def parse_sensors_sheet(self) -> pd.DataFrame:
        engine = 'xlrd' if self.is_xls else 'openpyxl'
        sheet_name = self._get_sheet_name('sensores')
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine, header=0)
        if 'Unnamed: 0' in df.columns or 'Exportación' in str(df.iloc[0, 0] if len(df) > 0 else ''):
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine, header=5)
        df = df.dropna(how='all').reset_index(drop=True)
        return df

    def _read_consumption_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Lee una hoja de consumos detectando automáticamente si tiene cabecera offset."""
        engine = 'xlrd' if self.is_xls else 'openpyxl'
        # GesMed y archivos de costes tienen 5 filas de metadatos antes de la cabecera
        if 'gesmed' in sheet_name.lower() or 'costes' in sheet_name.lower():
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine, header=5)
        else:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine, header=0)
        return df.dropna(how='all').reset_index(drop=True)

    def parse_consumptions_sheet(self) -> pd.DataFrame:
        """Parsea TODAS las hojas de tipo consumos y las combina en un único DataFrame."""
        sheets = self._find_all_sheets('consumos')
        if not sheets:
            raise ValueError("No se encontraron hojas de consumos")

        frames = []
        for sheet_name in sheets:
            try:
                df = self._read_consumption_sheet(sheet_name)
                df['_source_sheet'] = sheet_name  # trazabilidad del origen
                frames.append(df)
            except Exception as e:
                self.metadata[f"consumptions_error_{sheet_name}"] = str(e)

        if not frames:
            raise ValueError("No se pudo parsear ninguna hoja de consumos")

        combined = pd.concat(frames, ignore_index=True)
        return combined

    def parse_locations_sheet(self) -> pd.DataFrame:
        engine = 'xlrd' if self.is_xls else 'openpyxl'
        sheet_name = self._get_sheet_name('ubicaciones')
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine)
        df = df.dropna(how='all').reset_index(drop=True)
        return df

    def parse_all(self) -> Dict[str, pd.DataFrame]:
        result = {}
        try:
            result["sensores"] = self.parse_sensors_sheet()
            self.metadata["sensors_count"] = len(result["sensores"])
        except Exception as e:
            self.metadata["sensors_error"] = str(e)

        try:
            result["consumos"] = self.parse_consumptions_sheet()
            self.metadata["consumptions_count"] = len(result["consumos"])
        except Exception as e:
            self.metadata["consumptions_error"] = str(e)

        try:
            result["ubicaciones"] = self.parse_locations_sheet()
            self.metadata["locations_count"] = len(result["ubicaciones"])
        except Exception as e:
            self.metadata["locations_error"] = str(e)

        return result


SYSTEM_CODE_MAP = {
    'e11a1': 'Sistema Dalías (D)',
    'e11a4': 'Sistema Dalías (D)',
    'e11a5': 'Sistema Dalías (D)',
    'e12a1': 'Sistema Dalías (D)',
    'e12a5': 'Sistema Dalías (D)',
    'e12g1': 'Sistema Dalías (D)',
    'e12g2': 'Sistema Dalías (D)',
    'e12g5': 'Sistema Dalías (D)',
    'e12g6': 'Sistema Dalías (D)',
    'e12g7': 'Sistema Dalías (D)',
    'e12i1': 'Sistema Dalías (D)',
    'e12i3': 'Sistema Dalías (D)',
    'e12t1': 'Sistema Dalías (D)',
    'e1a1': 'Sistema Dalías (D)',
    'e52f1': 'Sistema Dalías (D)',
    'e21a1': 'Sistema Valdelentisco (D+D)',
    'e21i0': 'Sistema Valdelentisco (D+D)',
    'e22p0': 'Sistema Valdelentisco (D+D)',
    'e21b1': 'Sistema Torrevieja (D+D)',
    'e21b11': 'Sistema Torrevieja (D+D)',
    'e22f1': 'Sistema Torrevieja (D+D)',
    'e21e1': 'Sistema Águilas (D+D)',
    'e21e11': 'Sistema Águilas (D+D)',
    'e21e21': 'Sistema Águilas (D+D)',
    'e21e22': 'Sistema Águilas (D+D)',
    'e21e5': 'Sistema Águilas (D+D)',
    'e31b1': 'Sistema Mutxamel (D+D)',
    'e31b11': 'Sistema Mutxamel (D+D)',
    'e31e2': 'Sistema Oropesa (D+D)',
    'e31e21': 'Sistema Oropesa (D+D)',
    'e31e3': 'Sistema Moncófar (D+D)',
    'e31e31': 'Sistema Moncófar (D+D)',
    'e31f1': 'Sistema Oropesa (D+D)',
    'e31k0': 'Sistema Dalías (D)',
    'e31l1': 'Sistema Dalías (D)',
    'e32c1': 'Sistema Sagunto (D+D)',
    'e32d1': 'Sistema Sagunto (D+D)',
    'e32e1': 'Sistema Sagunto (D+D)',
    'e32j1': 'Sistema Sagunto (D+D)',
    'e32n6': 'Sistema Sagunto (D+D)',
    'e32n61': 'Sistema Sagunto (D+D)',
    'e32n7': 'Sistema Sagunto (D+D)',
    'e34b1': 'Sistema Sagunto (D+D)',
    'e34c1': 'Sistema Sagunto (D+D)',
    'e34c2': 'Sistema Sagunto (D+D)',
    'e34c3': 'Sistema Sagunto (D+D)',
    'e34c4': 'Sistema Sagunto (D+D)',
    'e34c5': 'Sistema Sagunto (D+D)',
    'e34c6': 'Sistema Sagunto (D+D)',
    'e34c7': 'Sistema Sagunto (D+D)',
    'e34c8': 'Sistema Sagunto (D+D)',
    'e34c9': 'Sistema Sagunto (D+D)',
    'e34c10': 'Sistema Sagunto (D+D)',
    'e34c11': 'Sistema Sagunto (D+D)',
    'e34c12': 'Sistema Sagunto (D+D)',
    'e6a1': 'Sistema Carboneras (D+D)',
    'e6a11': 'Sistema Carboneras (D+D)',
    'e6a3': 'Sistema Carboneras (D+D)',
    'e6a4': 'Sistema Dalías (D)',
    'e32n1': 'Sistema Sagunto (D+D)',
}


class CostesExplotacionCleaner:
    """Limpiador específico para archivos de costes de explotación"""
    
    def __init__(self, df: pd.DataFrame, file_upload_id: uuid.UUID, sheet_type: str):
        self.df = df.copy()
        self.file_upload_id = file_upload_id
        self.sheet_type = sheet_type  # 'consumos' o 'ubicaciones'
        self.errors: List[Dict[str, Any]] = []
        self.warnings: List[str] = []
        self.rows_processed = 0
    
    def clean(self) -> pd.DataFrame:
        if self.sheet_type == 'consumos':
            return self._clean_consumos()
        elif self.sheet_type == 'ubicaciones':
            return self._clean_ubicaciones()
        return self.df
    
    def _clean_consumos(self) -> pd.DataFrame:
        """Transforma datos de consumos al formato esperado por la BD"""
        # Convertir nombres de columnas a minúsculas
        self.df.columns = self.df.columns.str.lower().str.strip()
        
        # Mapeo de columnas del Excel a columnas esperadas
        column_mapping = {
            'nav_dim1_id': 'location_code',
            'subactuacion_id': 'subscriber_code',
            'subactuacion': 'consumption_type',
            'excel_descripcion': 'description',
            'real': 'volume_m3',
            'year': 'year',
            'mes': 'month'
        }
        
        # Renombrar columnas
        self.df = self.df.rename(columns=column_mapping)
        
        # Crear columna location_name usando el mapeo de códigos
        if 'location_code' in self.df.columns:
            self.df['location_name'] = (
                self.df['location_code']
                .astype(str)
                .str.strip()
                .str.lower()
                .map(SYSTEM_CODE_MAP)
            )
            unmapped = self.df['location_name'].isna().sum()
            if unmapped > 0:
                self.warnings.append(
                    f"{unmapped} registros con código de ubicación sin mapeo conocido"
                )
        
        # Crear columnas de fecha de período
        if 'year' in self.df.columns and 'month' in self.df.columns:
            self.df['period_start'] = pd.to_datetime(
                self.df['year'].astype(str) + '-' + 
                self.df['month'].astype(str).str.zfill(2) + '-01',
                errors='coerce'
            )
            # Calcular fin de mes
            self.df['period_end'] = self.df['period_start'] + pd.offsets.MonthEnd(0)
        
        # Asegurar que volume_m3 es numérico
        if 'volume_m3' in self.df.columns:
            self.df['volume_m3'] = pd.to_numeric(self.df['volume_m3'], errors='coerce')
        
        # Eliminar filas con datos incompletos
        required_cols = ['location_code', 'volume_m3', 'period_start']
        for col in required_cols:
            if col in self.df.columns:
                self.df = self.df.dropna(subset=[col])
        
        # Agregar tipo de consumo por defecto
        if 'consumption_type' not in self.df.columns:
            self.df['consumption_type'] = 'explotacion'
        else:
            # Limpiar y estandarizar tipo de consumo
            self.df['consumption_type'] = self.df['consumption_type'].fillna('explotacion')
            self.df['consumption_type'] = self.df['consumption_type'].str.lower().str.strip()
        
        # Agregar columna is_estimated
        self.df['is_estimated'] = False
        
        # Convertir Timestamps a strings para serialización JSON
        for col in ['period_start', 'period_end']:
            if col in self.df.columns:
                self.df[col] = self.df[col].astype(str)
        
        # Reemplazar NaN con None para serialización JSON
        self.df = self.df.where(pd.notnull(self.df), None)
        
        self.rows_processed = len(self.df)
        return self.df
    
    def _clean_ubicaciones(self) -> pd.DataFrame:
        """Transforma datos de ubicaciones al formato esperado por la BD"""
        # Convertir nombres de columnas a minúsculas
        self.df.columns = self.df.columns.str.lower().str.strip()
        
        # Mapeo de columnas del Excel a columnas esperadas
        column_mapping = {
            'inst. compuestas': 'name',
            'desaladoras': 'installation_type',
            'instalación': 'installation_subtype',
            'latitud': 'latitude',
            'longitud': 'longitude',
            'provincia': 'province',
            'zona': 'zone'
        }
        
        # Renombrar columnas
        self.df = self.df.rename(columns=column_mapping)
        
        # Asegurar que las coordenadas son numéricas
        for col in ['latitude', 'longitude']:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(self.df[col], errors='coerce')
        
        # Eliminar filas sin nombre
        if 'name' in self.df.columns:
            self.df = self.df.dropna(subset=['name'])
        
        # Usar province como region (la provincia ES la región en este contexto geográfico)
        if 'province' in self.df.columns:
            self.df['region'] = self.df['province']
        elif 'region' not in self.df.columns:
            self.df['region'] = 'Desconocida'
        
        self.rows_processed = len(self.df)
        return self.df
    
    def get_summary(self) -> Dict[str, Any]:
        return {
            "total_rows": len(self.df),
            "rows_processed": self.rows_processed,
            "errors": self.errors,
            "warnings": self.warnings,
            "columns": list(self.df.columns)
        }


class FileProcessor:
    def __init__(self, upload_dir: str = "/tmp/acuamed_uploads"):
        self.upload_dir = upload_dir
        os.makedirs(upload_dir, exist_ok=True)

    async def save_upload(self, file_content: bytes, filename: str) -> Tuple[str, int, str]:
        file_id = uuid.uuid4()
        extension = filename.split(".")[-1]
        stored_filename = f"{file_id}.{extension}"
        file_path = os.path.join(self.upload_dir, stored_filename)
        
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        file_size = os.path.getsize(file_path)
        file_hash = self._calculate_hash(file_content)
        
        return file_path, file_size, file_hash

    def _calculate_hash(self, content: bytes) -> str:
        return hashlib.sha256(content).hexdigest()

    def process_file(self, file_path: str, file_upload_id: uuid.UUID) -> Dict[str, Any]:
        parser = ExcelParser(file_path)
        valid, message = parser.validate_structure()
        
        if not valid:
            raise ExcelProcessingError(message)
        
        dataframes = parser.parse_all()
        results = {"metadata": parser.metadata}
        
        # Detectar si es un archivo de costes de explotación
        is_costes_explotacion = any(
            'gesmed' in str(name).lower() or 'costes' in str(name).lower() 
            for name in parser.metadata.get('sheets', [])
        )
        
        for sheet_name, df in dataframes.items():
            if df is not None and not df.empty:
                # Usar limpiador específico para archivos de costes de explotación
                if is_costes_explotacion and sheet_name in ['consumos', 'ubicaciones']:
                    cleaner = CostesExplotacionCleaner(df, file_upload_id, sheet_name)
                else:
                    cleaner = DataCleaner(df, file_upload_id)
                
                cleaned_df = cleaner.clean()
                results[sheet_name] = {
                    "data": sanitize_for_json(cleaned_df.to_dict("records")),
                    "summary": cleaner.get_summary()
                }
        
        return results

    def cleanup_file(self, file_path: str):
        if os.path.exists(file_path):
            os.remove(file_path)
